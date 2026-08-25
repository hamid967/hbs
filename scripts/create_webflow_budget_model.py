from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = Path('/home/ubuntu/hbs-hr/docs/Webflow_Development_Budget_Timeline.xlsx')
OUT.parent.mkdir(parents=True, exist_ok=True)

THEME = {'primary': '174D3C', 'light': 'E7F0EA', 'accent': 'B9924A', 'input': 'FFFDE7', 'success': 'E8F5E9', 'warning': 'FFF3E0', 'border': 'CBD5CF', 'text': '21382F'}
TITLE_FONT = Font(name='Arial', size=20, bold=True, color=THEME['primary'])
SUB_FONT = Font(name='Arial', size=10, italic=True, color='5E6C65')
HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
CELL_FONT = Font(name='Arial', size=10, color=THEME['text'])
SMALL_FONT = Font(name='Arial', size=9, color='5E6C65')
fill_primary = PatternFill('solid', fgColor=THEME['primary'])
fill_light = PatternFill('solid', fgColor=THEME['light'])
fill_input = PatternFill('solid', fgColor=THEME['input'])
fill_success = PatternFill('solid', fgColor=THEME['success'])
fill_warning = PatternFill('solid', fgColor=THEME['warning'])
thin = Side(style='thin', color=THEME['border'])
medium = Side(style='medium', color=THEME['primary'])

def setup(ws, title, subtitle, widths):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions['A'].width = 3
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.merge_cells('B2:H2')
    ws['B2'] = title
    ws['B2'].font = TITLE_FONT
    ws['B2'].alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[2].height = 34
    ws.merge_cells('B3:H3')
    ws['B3'] = subtitle
    ws['B3'].font = SUB_FONT
    ws['B3'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    ws.row_dimensions[3].height = 28

def header(ws, row, cols):
    for idx, value in enumerate(cols, start=2):
        c = ws.cell(row=row, column=idx, value=value)
        c.font = HEADER_FONT
        c.fill = fill_primary
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = Border(top=medium, bottom=medium, left=thin, right=thin)
    ws.row_dimensions[row].height = 28

def style_table(ws, start_row, end_row, start_col, end_col):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            cell.font = CELL_FONT
            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        ws.row_dimensions[r].height = 24

wb = Workbook()
overview = wb.active
overview.title = 'نظرة عامة'
assumptions = wb.create_sheet('الافتراضات')
budget = wb.create_sheet('تفاصيل الميزانية')
timeline = wb.create_sheet('الجدول الزمني')
risks = wb.create_sheet('المخاطر')

# Assumptions
setup(assumptions, 'افتراضات نموذج Webflow', 'الخلايا الصفراء قابلة للتعديل. القيم الافتراضية تخطيطية وليست عرض سعر من مورد.', {'B': 24, 'C': 18, 'D': 42, 'E': 18, 'F': 18, 'G': 18, 'H': 18})
header(assumptions, 5, ['فئة الفرضية', 'القيمة', 'الوحدة', 'ملاحظات'])
assumption_rows = [
    ['مدير المنتج / المشروع', 120, 'SAR/ساعة', 'سعر تخطيطي قابل للتعديل'],
    ['مصمم تجربة وواجهة', 115, 'SAR/ساعة', 'سعر تخطيطي قابل للتعديل'],
    ['مطور Webflow', 110, 'SAR/ساعة', 'سعر تخطيطي قابل للتعديل'],
    ['كاتب محتوى', 80, 'SAR/ساعة', 'سعر تخطيطي قابل للتعديل'],
    ['ضمان الجودة', 85, 'SAR/ساعة', 'سعر تخطيطي قابل للتعديل'],
    ['مهندس تكامل', 120, 'SAR/ساعة', 'سعر تخطيطي قابل للتعديل'],
    ['مختص SEO وقياس', 90, 'SAR/ساعة', 'سعر تخطيطي قابل للتعديل'],
    ['احتياطي مخاطر', 0.12, 'نسبة', 'يطبق على الإجمالي قبل الضريبة'],
    ['يوم العمل', 8, 'ساعة', 'لاحتساب الطاقة والمدة'],
    ['أسبوع العمل', 5, 'يوم', 'لاحتساب الأسابيع'],
    ['احتياطي مراجعة الإدارة', 4, 'يوم عمل', 'يضاف للمدة الأساسية'],
    ['اشتراك Webflow / استضافة', 0, 'SAR', 'أدخل عرض المورد أو التكلفة المعتمدة'],
    ['اسم النطاق والأصول المرخصة', 0, 'SAR', 'أدخل التكلفة الفعلية عند الاعتماد'],
    ['CRM أو تحليلات خارجية', 0, 'SAR', 'أدخل التكلفة الفعلية إن استخدمت'],
]
for r, values in enumerate(assumption_rows, start=6):
    for c, value in enumerate(values, start=2):
        assumptions.cell(r, c, value=value)
    assumptions.cell(r, 3).fill = fill_input
    assumptions.cell(r, 3).alignment = Alignment(horizontal='center', vertical='center')
    assumptions.cell(r, 3).number_format = '0.0%' if r == 13 else '#,##0.00'
style_table(assumptions, 6, 19, 2, 5)
assumptions.freeze_panes = 'B6'

# Budget Detail
setup(budget, 'تفاصيل ميزانية تنفيذ Webflow', 'يعتمد النموذج على ساعات العمل وسعر الساعة من ورقة الافتراضات. التكلفة الإجمالية تتغير عند تعديل ساعات العمل أو الأسعار أو الاحتياطي.', {'B': 16, 'C': 30, 'D': 20, 'E': 14, 'F': 12, 'G': 16, 'H': 16, 'I': 30})
header(budget, 5, ['المرحلة', 'حزمة العمل', 'الدور', 'سعر الساعة', 'الساعات', 'تكلفة العمل', 'تكلفة معدلة', 'ملاحظة'])
work = [
    ('1. الحوكمة والنطاق', 'نطاق، ملاك محتوى، سياسة نشر', 'مدير المنتج / المشروع', 16, 'اعتماد حدود Webflow والتطبيق'),
    ('1. الحوكمة والنطاق', 'نطاق، ملاك محتوى، سياسة نشر', 'مصمم تجربة وواجهة', 8, 'مراجعة الاتجاه والهوية'),
    ('1. الحوكمة والنطاق', 'نطاق، ملاك محتوى، سياسة نشر', 'مطور Webflow', 4, 'تقدير هيكل التنفيذ'),
    ('2. نظام التصميم', 'Variables ومكونات RTL وإتاحة', 'مصمم تجربة وواجهة', 24, 'مكونات قابلة لإعادة الاستخدام'),
    ('2. نظام التصميم', 'Variables ومكونات RTL وإتاحة', 'مطور Webflow', 16, 'إنشاء المكونات والأنماط'),
    ('2. نظام التصميم', 'Variables ومكونات RTL وإتاحة', 'مدير المنتج / المشروع', 8, 'مراجعة معايير القبول'),
    ('3. الصفحات الأساسية', 'الرئيسية والمنصة والحلول والميزات', 'مصمم تجربة وواجهة', 32, 'تصميم الصفحات ومسارات CTA'),
    ('3. الصفحات الأساسية', 'الرئيسية والمنصة والحلول والميزات', 'مطور Webflow', 48, 'بناء الصفحات المتجاوبة'),
    ('3. الصفحات الأساسية', 'الرئيسية والمنصة والحلول والميزات', 'كاتب محتوى', 24, 'نسخ عربي وتسميات'),
    ('3. الصفحات الأساسية', 'الرئيسية والمنصة والحلول والميزات', 'ضمان الجودة', 8, 'فحص الهاتف والتنقل'),
    ('3. الصفحات الأساسية', 'الرئيسية والمنصة والحلول والميزات', 'مدير المنتج / المشروع', 16, 'توجيه ومراجعة'),
    ('4. CMS والمحتوى', 'مجموعات وقوالب محتوى', 'مطور Webflow', 24, 'إنشاء CMS والقوالب'),
    ('4. CMS والمحتوى', 'مجموعات وقوالب محتوى', 'كاتب محتوى', 32, 'إعداد المحتوى الأولي'),
    ('4. CMS والمحتوى', 'مجموعات وقوالب محتوى', 'ضمان الجودة', 12, 'اختبار القوالب والتحرير'),
    ('4. CMS والمحتوى', 'مجموعات وقوالب محتوى', 'مدير المنتج / المشروع', 8, 'حوكمة المحتوى'),
    ('5. التحويل والتكامل', 'نموذج طلب العرض وتدفق CRM اختياري', 'مطور Webflow', 20, 'نموذج وتحويلات'),
    ('5. التحويل والتكامل', 'نموذج طلب العرض وتدفق CRM اختياري', 'مهندس تكامل', 16, 'Webhook أو CRM بعد الاعتماد'),
    ('5. التحويل والتكامل', 'نموذج طلب العرض وتدفق CRM اختياري', 'ضمان الجودة', 8, 'اختبارات الإرسال والأخطاء'),
    ('5. التحويل والتكامل', 'نموذج طلب العرض وتدفق CRM اختياري', 'مدير المنتج / المشروع', 8, 'اعتماد الحقول والخصوصية'),
    ('6. SEO والجودة والإطلاق', 'SEO، قياس، قبول، إطلاق', 'مختص SEO وقياس', 16, 'عناوين ووصف وأحداث'),
    ('6. SEO والجودة والإطلاق', 'SEO، قياس، قبول، إطلاق', 'مطور Webflow', 12, 'تحسينات ونشر'),
    ('6. SEO والجودة والإطلاق', 'SEO، قياس، قبول، إطلاق', 'ضمان الجودة', 20, 'قبول صفحات ونماذج'),
    ('6. SEO والجودة والإطلاق', 'SEO، قياس، قبول، إطلاق', 'مدير المنتج / المشروع', 8, 'بوابة الإطلاق'),
]
rate_cells = {
    'مدير المنتج / المشروع': '$C$6', 'مصمم تجربة وواجهة': '$C$7', 'مطور Webflow': '$C$8',
    'كاتب محتوى': '$C$9', 'ضمان الجودة': '$C$10', 'مهندس تكامل': '$C$11', 'مختص SEO وقياس': '$C$12'
}
for r, (phase, package, role, hours, note) in enumerate(work, start=6):
    budget.cell(r, 2, phase); budget.cell(r, 3, package); budget.cell(r, 4, role)
    budget.cell(r, 5, f'=الافتراضات!{rate_cells[role]}')
    budget.cell(r, 6, hours)
    budget.cell(r, 7, f'=E{r}*F{r}')
    budget.cell(r, 8, f'=G{r}')
    budget.cell(r, 9, note)
style_table(budget, 6, 5 + len(work), 2, 9)
for r in range(6, 6 + len(work)):
    for c in (5, 7, 8): budget.cell(r, c).number_format = '#,##0.00 [$SAR]'
    budget.cell(r, 6).number_format = '#,##0.0'
last = 5 + len(work)
budget.cell(last + 2, 2, 'إجمالي تكلفة العمل')
budget.cell(last + 2, 7, f'=SUM(G6:G{last})')
budget.cell(last + 3, 2, 'تكاليف خارجية قابلة للإدخال')
budget.cell(last + 3, 7, '=SUM(الافتراضات!C17:C19)')
budget.cell(last + 4, 2, 'احتياطي المخاطر')
budget.cell(last + 4, 7, f'=(G{last+2}+G{last+3})*الافتراضات!C13')
budget.cell(last + 5, 2, 'الإجمالي التقديري قبل الضريبة')
budget.cell(last + 5, 7, f'=SUM(G{last+2}:G{last+4})')
for r in range(last+2, last+6):
    budget.cell(r, 2).font = Font(name='Arial', bold=True, color=THEME['text'])
    budget.cell(r, 7).font = Font(name='Arial', bold=True, color=THEME['text'])
    budget.cell(r, 7).number_format = '#,##0.00 [$SAR]'
    for c in range(2, 8):
        budget.cell(r, c).fill = fill_success if r == last+5 else fill_light
        budget.cell(r, c).border = Border(top=thin, bottom=thin, left=thin, right=thin)
budget.freeze_panes = 'B6'; budget.auto_filter.ref = f'B5:I{last}'

# Timeline
setup(timeline, 'الجدول الزمني التقديري لتنفيذ Webflow', 'المدة باليوم التشغيلي. لا تبدأ أي تكاملات خارجية قبل اعتماد المزوّد والحقول وسياسة الخصوصية.', {'B': 18, 'C': 34, 'D': 16, 'E': 16, 'F': 16, 'G': 28, 'H': 28})
header(timeline, 5, ['المرحلة', 'المخرج', 'المدة (أيام عمل)', 'البداية', 'النهاية', 'الاعتمادية', 'بوابة القبول'])
timeline_rows = [
    ('1. الحوكمة والنطاق', 'خريطة نطاق ومالك محتوى وسياسة نشر', 3, 'بداية البرنامج', 'فصل Webflow عن التطبيق الداخلي'),
    ('2. نظام التصميم', 'مكونات RTL ومتغيرات وإتاحة', 5, 'اعتماد النطاق', 'تباين وتركيز وحركة مختزلة'),
    ('3. الصفحات الأساسية', 'الرئيسية والمنصة والحلول والميزات', 7, 'نظام التصميم', 'CTA ومسارات الهاتف صحيحة'),
    ('4. CMS والمحتوى', 'مجموعات CMS وقوالب تحرير', 4, 'الصفحات الأساسية', 'مالك وتاريخ مراجعة لكل عنصر'),
    ('5. التحويل والتكامل', 'نموذج طلب العرض وتدفق CRM اختياري', 4, 'سياسة الخصوصية وتحديد CRM', 'لا مفاتيح في الواجهة'),
    ('6. SEO والجودة والإطلاق', 'SEO وقياس وقبول ونشر', 3, 'اكتمال الصفحات والنموذج', 'لا عوائق حرجة'),
    ('مراجعة الإدارة', 'مراجعات محتوى وقرار نشر', '=الافتراضات!C16', 'إتمام العمل', 'محضر قرار النشر'),
]
for r, row in enumerate(timeline_rows, start=6):
    phase, output, duration, dep, gate = row
    timeline.cell(r, 2, phase); timeline.cell(r, 3, output); timeline.cell(r, 4, duration)
    timeline.cell(r, 5, '=DATE(2026,1,1)' if r == 6 else f'=F{r-1}+1')
    timeline.cell(r, 6, f'=E{r}+D{r}-1')
    timeline.cell(r, 7, dep); timeline.cell(r, 8, gate)
style_table(timeline, 6, 12, 2, 8)
for r in range(6, 13):
    timeline.cell(r, 4).number_format = '#,##0.0'
    timeline.cell(r, 5).number_format = 'dd-mmm-yyyy'; timeline.cell(r, 6).number_format = 'dd-mmm-yyyy'
timeline.cell(14, 2, 'إجمالي مدة التنفيذ')
timeline.cell(14, 4, '=SUM(D6:D12)')
timeline.cell(15, 2, 'الأسابيع التقديرية')
timeline.cell(15, 4, '=D14/الافتراضات!C15')
for r in (14,15):
    for c in range(2,5):
        timeline.cell(r,c).fill = fill_success; timeline.cell(r,c).border = Border(top=thin,bottom=thin,left=thin,right=thin)
        timeline.cell(r,c).font = Font(name='Arial', bold=True, color=THEME['text'])
timeline.cell(15,4).number_format = '0.0'
timeline.freeze_panes='B6'

# Risks
setup(risks, 'مخاطر الميزانية والجدول الزمني', 'يسجل هذا السجل ما قد يغير التكلفة أو المدة؛ يُحدّث في مراجعة أسبوعية.', {'B': 26, 'C': 14, 'D': 14, 'E': 14, 'F': 35, 'G': 20, 'H': 20})
header(risks, 5, ['الخطر', 'الاحتمال', 'الأثر', 'الأولوية', 'التحكم', 'المالك', 'مؤشر الإنذار'])
risk_rows = [
    ('توسع نطاق الصفحات أو المحتوى بعد اعتماد التصميم', 'متوسط', 'مرتفع', 'P1', 'سجل تغيير وموافقة مالك المنتج قبل البدء', 'مالك المنتج', 'طلبات إضافية بلا نطاق أو تقدير'),
    ('تأخر اعتماد المحتوى أو الملاك', 'مرتفع', 'متوسط', 'P1', 'مالك وتاريخ مراجعة لكل عنصر CMS', 'التسويق', 'عنصر محتوى بلا مراجع'),
    ('تأخر اختيار CRM أو تكامل خارجي', 'متوسط', 'مرتفع', 'P1', 'فصل مسار النموذج عن التكامل الخارجي', 'مالك القرار', 'مزوّد أو حقول غير معتمدة'),
    ('تكلفة مزود أو أصل مرخص غير محسوبة', 'متوسط', 'متوسط', 'P2', 'إدخال عروض الموردين في الافتراضات قبل الالتزام', 'المالية/التسويق', 'تكلفة خارجية = صفر مع حاجة مثبتة'),
    ('تراجع جودة RTL أو الإتاحة مع تسارع النشر', 'منخفض', 'مرتفع', 'P2', 'بوابة جودة للهاتف والتركيز والنموذج', 'الجودة', 'فشل اختبار تنقل أو تباين'),
]
for r, values in enumerate(risk_rows, start=6):
    for c, value in enumerate(values, start=2): risks.cell(r,c,value=value)
style_table(risks, 6, 10, 2, 8)
risks.freeze_panes='B6'; risks.auto_filter.ref='B5:H10'

# Overview after references are ready
setup(overview, 'نموذج ميزانية وجدول Webflow', 'نموذج تخطيطي قابل للتعديل بالريال السعودي. عدل الأسعار والساعات والتكاليف الخارجية في ورقة الافتراضات؛ يعاد حساب النموذج عند فتحه في Excel.', {'B': 32, 'C': 18, 'D': 18, 'E': 18, 'F': 18, 'G': 18, 'H': 18})
overview.merge_cells('B5:H5'); overview['B5']='ملخص تنفيذي'; overview['B5'].font=Font(name='Arial',size=14,bold=True,color=THEME['primary']); overview['B5'].fill=fill_light; overview['B5'].alignment=Alignment(horizontal='right')
summary = [
    ('تكلفة العمل الأساسية', f"='تفاصيل الميزانية'!G{last+2}", 'SAR'),
    ('التكاليف الخارجية المدخلة', f"='تفاصيل الميزانية'!G{last+3}", 'SAR'),
    ('احتياطي المخاطر', f"='تفاصيل الميزانية'!G{last+4}", 'SAR'),
    ('الإجمالي التقديري قبل الضريبة', f"='تفاصيل الميزانية'!G{last+5}", 'SAR'),
    ('مدة التنفيذ', "='الجدول الزمني'!D14", 'يوم عمل'),
    ('المدة بالأسابيع', "='الجدول الزمني'!D15", 'أسبوع'),
]
header(overview, 6, ['المؤشر', 'القيمة المحسوبة', 'الوحدة'])
for r, (label, formula, unit) in enumerate(summary, start=7):
    overview.cell(r,2,label); overview.cell(r,3,formula); overview.cell(r,4,unit)
style_table(overview, 7, 12, 2, 4)
for r in range(7,11): overview.cell(r,3).number_format='#,##0.00 [$SAR]'
overview.cell(12,3).number_format='0.0'
for c in range(2,5): overview.cell(10,c).fill=fill_success
overview.merge_cells('B15:H15'); overview['B15']='سيناريوهات التخطيط'; overview['B15'].font=Font(name='Arial',size=14,bold=True,color=THEME['primary']); overview['B15'].fill=fill_light
header(overview,16,['السيناريو','معامل التكلفة','معامل المدة','تكلفة مقدرة','أيام عمل','ملاحظة'])
scenarios=[('مقتصد',0.85,0.90,'نطاق ثابت ومحتوى جاهز'),('أساسي',1.00,1.00,'الافتراضات الحالية'),('موسع',1.20,1.20,'دورات مراجعة أو تكامل أعلى')]
for r,(name,cost_factor,time_factor,note) in enumerate(scenarios,start=17):
    overview.cell(r,2,name); overview.cell(r,3,cost_factor); overview.cell(r,4,time_factor)
    overview.cell(r,5,f'=C{r}*$C$10'); overview.cell(r,6,f'=D{r}*$C$11'); overview.cell(r,7,note)
style_table(overview,17,19,2,7)
for r in range(17,20):
    overview.cell(r,3).number_format='0%'; overview.cell(r,4).number_format='0%'; overview.cell(r,5).number_format='#,##0.00 [$SAR]'; overview.cell(r,6).number_format='0.0'
overview.merge_cells('B22:H22'); overview['B22']='روابط الأوراق'; overview['B22'].font=Font(name='Arial',size=14,bold=True,color=THEME['primary']); overview['B22'].fill=fill_light
for i,sheet in enumerate(['الافتراضات','تفاصيل الميزانية','الجدول الزمني','المخاطر'],start=23):
    cell=overview.cell(i,2,sheet); cell.hyperlink=f"#'{sheet}'!A1"; cell.font=Font(name='Arial',size=10,color=THEME['primary'],underline='single')
overview.merge_cells('B29:H30'); overview['B29']='تنبيه: النموذج تقديري للتخطيط. لا يشمل ضريبة القيمة المضافة أو رسوم الموردين غير المدخلة. اعتمد أسعار الساعة وتكاليف Webflow وCRM والأصول من عروض موثقة قبل الالتزام المالي.'; overview['B29'].font=SMALL_FONT; overview['B29'].alignment=Alignment(horizontal='right',vertical='center',wrap_text=True); overview['B29'].fill=fill_warning
overview.freeze_panes='B6'

# Chart
chart = BarChart(); chart.title = 'تكلفة العمل حسب المرحلة'; chart.y_axis.title = 'SAR'; chart.x_axis.title = 'المرحلة'
phase_summary_start = last + 8
budget.cell(phase_summary_start,2,'ملخص تكلفة العمل حسب المرحلة'); budget.cell(phase_summary_start,2).font=Font(name='Arial',bold=True,color=THEME['primary'])
header(budget, phase_summary_start+1, ['المرحلة','التكلفة'])
phases = ['1. الحوكمة والنطاق','2. نظام التصميم','3. الصفحات الأساسية','4. CMS والمحتوى','5. التحويل والتكامل','6. SEO والجودة والإطلاق']
for idx, phase in enumerate(phases, start=phase_summary_start+2):
    budget.cell(idx,2,phase); budget.cell(idx,3,f'=SUMIF(B6:B{last},B{idx},G6:G{last})'); budget.cell(idx,3).number_format='#,##0.00 [$SAR]'
style_table(budget,phase_summary_start+2,phase_summary_start+7,2,3)
data=Reference(budget,min_col=3,min_row=phase_summary_start+1,max_row=phase_summary_start+7)
cats=Reference(budget,min_col=2,min_row=phase_summary_start+2,max_row=phase_summary_start+7)
chart.add_data(data,titles_from_data=True); chart.set_categories(cats); chart.height=7; chart.width=15
chart.series[0].graphicalProperties.solidFill=THEME['primary']
budget.add_chart(chart,'K6')
budget.conditional_formatting.add(f'G6:G{last}', DataBarRule(start_type='min', end_type='max', color=THEME['accent']))

for ws in wb.worksheets:
    ws.sheet_properties.pageSetUpPr.fitToPage=True
    ws.page_setup.fitToWidth=1
    ws.page_setup.fitToHeight=0
    ws.sheet_view.rightToLeft=True

wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.save(OUT)
print(OUT)

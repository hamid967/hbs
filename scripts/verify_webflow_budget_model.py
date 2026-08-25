from pathlib import Path
from openpyxl import load_workbook

path = Path('/home/ubuntu/hbs-hr/docs/Webflow_Development_Budget_Timeline.xlsx')
if not path.exists():
    raise SystemExit('Budget workbook was not created')

wb = load_workbook(path, data_only=False)
expected = ['نظرة عامة', 'الافتراضات', 'تفاصيل الميزانية', 'الجدول الزمني', 'المخاطر']
if wb.sheetnames != expected:
    raise SystemExit(f'Unexpected worksheets: {wb.sheetnames}')

checks = {
    ('تفاصيل الميزانية', 'G30'): '=SUM(G6:G28)',
    ('تفاصيل الميزانية', 'G33'): '=SUM(G30:G32)',
    ('الجدول الزمني', 'D14'): '=SUM(D6:D12)',
    ('الجدول الزمني', 'D15'): '=D14/الافتراضات!C15',
    ('نظرة عامة', 'C10'): "='تفاصيل الميزانية'!G33",
    ('نظرة عامة', 'E17'): '=C17*$C$10',
}
for (sheet, cell), expected_formula in checks.items():
    actual = wb[sheet][cell].value
    if actual != expected_formula:
        raise SystemExit(f'{sheet}!{cell}: expected {expected_formula}, got {actual}')

for sheet in expected:
    if not wb[sheet].sheet_view.rightToLeft:
        raise SystemExit(f'{sheet} is not configured for RTL')

print('Webflow budget workbook verification passed')
